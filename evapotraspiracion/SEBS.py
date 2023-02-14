from pythonProject.heat_fluxes_italo import KB_1, z0h, Cw, PSIh_y, u_pbl, FRUstar
import Formule_virginia as fv
import numpy as np
import openpyxl as xl
import rasterio as rio
from rasterio.plot import show
from metpy.units import units
import pandas as pd
import scipy

###~~CARICO I DATI~~###################################################################################

#LST
my_lst = rio.open(r"C:\Users\1\Desktop\20180718_1155_LST_EMISIVIDAD_ALBEDO_NDVI\1155LST.gcu")
show(my_lst)

my_lse = rio.open(r"C:\Users\1\Desktop\20180718_1155_LST_EMISIVIDAD_ALBEDO_NDVI\1155albedo.gcu")
show(my_lse)

with rio.open(r"C:\Users\1\Desktop\20180718_1155_LST_EMISIVIDAD_ALBEDO_NDVI\1155LST.gcu") as lst_raster:
    lst_b1 = lst_raster.read(1, masked=False)
    profile = lst_raster.profile
    lst_b1[lst_b1 == 0] = [np.nan]
    lst_b1 = lst_b1#[329:526,50:51]

#carico il wb
wb = xl.load_workbook(r"C:\Users\1\Desktop\fluxes_Braccagni\meteo_cornJUN_JUL1.xlsx")
sheet = wb['meteo_cornJUN_JUL']

#Tref
my_t_ref = sheet.cell(row=13681, column=15) #questo e il valore che mi serve, Twet4 alle 11.55 del 18.07.2018
my_t_ref = pd.to_numeric(my_t_ref.value, errors='coerce', downcast='float')
my_t_ref_unit = units.Quantity(my_t_ref, "°C")
print(my_t_ref)
t_ref_array = np.full_like(lst_b1, my_t_ref)

#carico l'immagine ndvi
my_ndvi = rio.open(r"C:\Users\1\Desktop\20180718_1155_LST_EMISIVIDAD_ALBEDO_NDVI\1155NDVI.gcu")
show(my_ndvi)

#calcolo fvc
with rio.open(r"C:\Users\1\Desktop\20180718_1155_LST_EMISIVIDAD_ALBEDO_NDVI\1155NDVI.gcu") as ndvi_raster:
    ndvi_b1 = ndvi_raster.read(1, masked=False)
    ndvi_b1[ndvi_b1 < 0] = [np.nan]
    ndvi_b1 = ndvi_b1

my_fvc = ndvi_b1 - 0.05

with rio.open("fvc_calc.gcu", 'w', **profile) as fvc_raster:
    fvc_raster.write(my_fvc, 1, )
show(my_fvc)

#calcolo il lai, in tre modi differenti e vedo come cambia l'ET
with rio.open("fvc_calc.gcu") as fvc_raster:
    fvc_b1 = fvc_raster.read(1, masked=False)
    profile = fvc_raster.profile
    fvc_b1[fvc_b1 < 0] = [0]
    fvc_b1 = fvc_b1

my_lai_fan = (+1.424*ndvi_b1 -0.0897) #Fan et al 2008 (grasslands)
my_lai = np.sqrt(ndvi_b1 * (1.0 + ndvi_b1) / (1.0 - ndvi_b1 + 1.0E-6))
Kpar = 0.5  # vedi Fisher 2008 e Ross 1976
my_lai_fvc = (-np.log(1-fvc_b1/Kpar))

with rio.open("lai_calc.gcu", 'w', **profile) as lai_raster:
    lai_raster.write(my_lai, 1, )
show(my_lai)


my_z0m = np.exp(-5.5 + 5.8*ndvi_b1) #Tomelloso method, Bolle and Streckbach 1993
#my_z0m = 0.018*lai_b1 #brutsaert et al 1982
#c'è anche un altro metodo più complicato, in cui devo calcolare io delle valriabili sulla base dei pixel, che potrebbe essere una buona soluzione
with rio.open("z0m_calc.gcu", 'w', **profile) as z0m_raster:
    z0m_raster.write(my_z0m, 1, )
#show(my_z0m)

#laiB1
with rio.open("lai_calc.gcu") as lai_raster:
    lai_b1 = lai_raster.read(1, masked=False)

#calcolo hc da z0m, z0m1 e z0m2
with rio.open("z0m_calc.gcu") as z0m_raster:
    z0m_b1 = z0m_raster.read(1, masked=False)

#my_hc = z0m_b1/0.12
#my_hc = 0.15 * lai_b1 #SEBAL manual
my_hc = z0m_b1/0.136 #Brutsaert, 1982
#più o meno è uguale

#Alfalfa: H = 0.009LAI^2 + 0.076LAI
#Corn: H = 0.03LAI^3 - 0.2194LAI^2 + 0.7243LAI

with rio.open("hc_calc.gcu", 'w', **profile) as hc_raster:
    hc_raster.write(my_hc, 1, )
show(my_hc)

#calcolo w_ref da excel
my_w_ref = sheet.cell(row=13681, column=21) #questo e il valore che mi serve, wind4 alle 11.55 del 18.07.2018 (m/s)
my_w_ref = pd.to_numeric(my_w_ref.value, errors='coerce', downcast='float')
w_ref_array = np.full_like(lst_b1, my_w_ref)

#calcolo ea_ref
tair_dry4 = sheet.cell(row=13681, column=14) #ore 11.55 del 18.07.2018 (Celsius)
tair_dry4.value = pd.to_numeric(tair_dry4.value, errors='coerce', downcast='float')
tair_dry4 = units.Quantity(tair_dry4.value, "°C")
#tair_wet4 = my_t_ref.value
my_press_ref = 100858.1117 #Pa, valure ottenuto da ERA5
my_press_ref = units.Quantity(my_press_ref, "Pa")

press_ref_array = np.full_like(lst_b1, my_press_ref)

my_ea_ref = fv.psychrometric_vapor_pressure_wet(my_press_ref, tair_dry4, my_t_ref_unit, psychrometer_coefficient=None)
dew = fv.dewpoint(my_ea_ref)
my_q_ref = fv.specific_humidity_from_dewpoint(my_press_ref, dew)    #dimensionless
my_q_ref = units.Quantity(my_q_ref.magnitude, "g/kg")
q_ref_array = np.full_like(lst_b1, my_q_ref)


#calcolo SWnet e LWnet
my_SWin = sheet.cell(row=13681, column=33) #questo e il valore che mi serve, SWIN alle 11.55 del 18.07.2018 (m/s)
my_SWin.value = pd.to_numeric(my_SWin.value, errors='coerce', downcast='float')
print(my_SWin.value)
my_LWin = sheet.cell(row=13681, column=35) #questo e il valore che mi serve, LWIN alle 11.55 del 18.07.2018 (m/s)
my_LWin.value = pd.to_numeric(my_LWin.value, errors='coerce', downcast='float')
print(my_LWin.value)
sigma = 5.6704 * 10 ** -8  # boltzmann constant
B_constant = sigma


#emissività, nel modello metto questa e non quella calcolata con la formula
with rio.open(r"C:\Users\1\Desktop\20180718_1155_LST_EMISIVIDAD_ALBEDO_NDVI\1155LSE.gcu") as lse_raster:
    lse_b1 = lse_raster.read(1, masked=False)
    lse_b1[lse_b1 == 0] = [np.nan]
    lse_b1 = lse_b1

with rio.open(r"C:\Users\1\Desktop\20180718_1155_LST_EMISIVIDAD_ALBEDO_NDVI\1155albedo.gcu") as albedo_raster:
    albedo_b1 = albedo_raster.read(1, masked=False)
    albedo_b1[albedo_b1 == 0] = [np.nan]
    albedo_b1 = albedo_b1

my_SWnet = (1-albedo_b1)*my_SWin.value

with rio.open("SWnet_calc.gcu", 'w', **profile) as SWnet_raster:
    SWnet_raster.write(my_SWnet, 1, )
    profile = SWnet_raster.profile
show(my_SWnet)

my_LWnet = (lse_b1*my_LWin.value) - (lse_b1*B_constant*(lst_b1**4))

with rio.open("LWnet_calc.gcu", 'w', **profile) as LWnet_raster:
    LWnet_raster.write(my_LWnet, 1, )
    profile = LWnet_raster.profile
show(my_LWnet)

#saco la press_surf
my_press_surf = sheet.cell(row=13681, column=25) #pressure alle 11.55 del 18.07.2018 (kPa) ATTENZIONE QUI, MI DA COME VALORE 1013.7!
my_press_surf = pd.to_numeric(my_press_surf.value, errors='coerce', downcast='float')
my_press_surf = my_press_surf*100 #Pa
print(my_press_surf)

press_surf_array = np.full_like(lst_b1, my_press_surf)

#z_ref
my_z_ref = 2 #(m)
z_ref_array = np.full_like(lst_b1, my_z_ref)


my_Rn = my_SWnet + my_LWnet
my_G0 = my_Rn * (0.05 + (1 - fvc_b1) * (0.315 - 0.05)) #soil heat flux [W/m2]


#my_G0 potrebbe essere calcolato anche così (la funzione è da sistemare) # sebal manual
#my_G0 = (my_lst/my_lse(*(0.0038*my_lse + 0.0074*my_lse**2)(1 - 0.98*my_ndvi**4)))*my_Rn


########### rinomino le variabili di input

RN = my_Rn  # insert variable - Net Radiation [W/m2]
Ta = t_ref_array  # insert variable - Air Temperature [ÂªC]
LST = lst_b1  # insert variable Land Surface Temperature [K]
LAI = my_lai  # insert variable LAI [m2/m2]
RH = q_ref_array  # Relative Humidity [%]
u = w_ref_array  # Wind speed at 2m [m/s]
TaK = Ta + 273.16  # Air Temperature [K]
NDVI = ndvi_b1  # insert - variable NDVI [-]
z0m = my_z0m
fvc = my_fvc
press_surf = press_surf_array


alt_ms = z_ref_array  # [m] Height of measurement
z_pbl = np.full_like(lst_b1, 1000)  # creo un array con valore 1000, [m] Height of Planetary Boundary Layer
hst = 0.12 * z_pbl  # Height of Atmospheric Boundary Layer Similary Stull.1988
z_ref = z_ref_array  # heightness reference
p_s = press_surf_array  # Pa
ro = 1.23  # [kg/m3] dry air density
Cp = 1004  # [JK/kg] air specific heat
# DEFINITION OF ASL HEIGHT
alfa = 0.12  # Medium value of alpha defined by Brutsaert (1999)
beta = 125.0  # Medium value of beta defined by Brutsaert (1999)
hbpl = alfa * (beta / alfa)  # Height of Atmospheric Boundary Layer (ABL)
Cd = 0.2  # Drag coefficient of the foliage elements
Rv = 461.05  # specific gas constant water vapour (J kg-1 K-1)
Rd = 287.04  # specific gas constant dry air (J kg-1 K-1)
k = 0.41  # Von Karmann Constant
g = 9.81  # Gravity force [m/s2]
Alt = 3 # 299  # Altitude [m]
DOY = 199 #200  # DAY OF THE YEAR
Cdi = -7 * (1e-06) * ((DOY)) ** 2 + 0.0027 * (DOY) + 0.124

sigma = 5.6704 * 10 ** -8  # boltzmann constant
es = 6.122 * np.exp((17.67 * (TaK - 273.16)) / ((TaK - 273.16) + 243.5))  # [HPa]
ea = (RH / 100) * es  # my_ea??? posso lasciarlo così [HPa]
Ea = 1.24 * (ea / TaK) ** (1. / 7.)  # air emissivity


h = my_hc #1  # canopy height [1m]

############################### dati italo ##################################
# Rx = 100# insert variable - Shortwave Radiation [W/m2]
#RN = np.array([670])  # insert variable - Net Radiation [W/m2]
#Ta = np.array([30])  # insert variable - Air Temperature [ÂªC]
#LST = np.array([320])  # insert variable Land Surface Temperature [K]
#LAI = np.array([1])  # insert variable LAI [m2/m2]
#RH = np.array([99])  # Relative Humidity [%]
#u = np.array([2])  # Wind speed at 2m [m/s]
#TaK = Ta + 273.16  # Air Temperature [K]
#NDVI = np.array([0.40])  # insert - variable NDVI [-]

#z0m = 0.018 * LAI

#alt_ms = 2  # [m] Height of measurement
#z_pbl = 1000  # [m] Height of Planetary Boundary Layer
#hst = 0.12 * z_pbl  # Height of Atmospheric Boundary Layer Similary Stull.1988
#z_ref = 2  # heightness reference
#p_s = 101325  # Pa
#ro = 1.23  # [kg/m3] dry air density
#Cp = 1004  # [JK/kg] air specific heat
# DEFINITION OF ASL HEIGHT
#alfa = 0.12  # Medium value of alpha defined by Brutsaert (1999)
#beta = 125.0  # Medium value of beta defined by Brutsaert (1999)
#hbpl = alfa * (beta / alfa)  # Height of Atmospheric Boundary Layer (ABL)
#Cd = 0.2  # Drag coefficient of the foliage elements
#Rv = 461.05  # specific gas constant water vapour (J kg-1 K-1)
#Rd = 287.04  # specific gas constant dry air (J kg-1 K-1)
#k = 0.41  # Von Karmann Constant
#g = 9.81  # Gravity force [m/s2]
#Alt = 299  # Altitude [m]
#DOY = 200  # DAY OF THE YEAR
#Cdi = -7 * (1e-06) * ((DOY)) ** 2 + 0.0027 * (DOY) + 0.124

#sigma = 5.6704 * 10 ** -8  # boltzmann constant
#es = 6.122 * np.exp((17.67 * (TaK - 273.16)) / ((TaK - 273.16) + 243.5))  # [HPa]
#ea = (RH / 100) * es  # [HPa]
#Ea = 1.24 * (ea / TaK) ** (1. / 7.)  # air emissivity

#h = 1  # canopy height [1m]

###~~DA QUI NON HO NUOVE VARIABILI, È SOLO CALCOLO~~###################
d0 = 0.67 * h
z0 = 0.136 * h  # Brutsaert (1982)
pa = 101.3 * ((293 - 0.0065 * Alt) / 293) ** 5.26  # [Kpa]
es_kpa = es / 10  # [KPa]
ea_kpa = ea / 10  # [KPa]
ea_pa = ea * 100
ee = 0.622  # Water Vapor/Dry Air mix ratio
rv = (ee * es_kpa) / (pa - es_kpa)  # Mixing Ratio
Tv = TaK * (1 + (rv / ee)) / (1 + rv)  # Wallace, John M.; Hobbs, Peter V. (2006). Atmospheric Science. ISBN 0-12-732951-X.
ps = p_s * ((44331.0 - (Alt + z_pbl)) / (44331.0 - alt_ms)) ** (1.0 / 0.1903)
rhoam = (ps / (Rd * LST)) * (1.0 - 0.378 * ea_pa / ps)  # moist air density (kg m-3)
d = 2.0 / 3.0 * h  # zero plane displacement (m)
# FRICTION VELOCITY IN CONDITION OF STABILITY
uFi = (k * u) / np.log(2 / z0m)
nec = (Cd * LAI) / ((2 * (uFi ** 2)) / u ** 2)  # Within-canopy wind profile extinction coefficient
NDVI_FC = np.copy(NDVI)
NDVI_FC[NDVI_FC < 0.2] = 0.2
NDVI_FC[NDVI_FC > 0.8] = 0.8
#Fc = ((NDVI_FC - 0.2) / (0.8 - 0.2)) ** 2  # Fraction of Vegetation Cover (Sobrino, 2008)
Fc = fvc #uso quello calcolato da me
# Fs = 1-Fc  #Fracction of Soil Cover
kB = KB_1(u, uFi, z0m, Fc, LAI, TaK, pa, h)
show(kB)
Z0h = z0h(kB, z0m)
show(Z0h)
Z0h[Z0h == np.inf] = 28814991000.0  # 375579840.0 #high resistance
Z0h[Z0h > 3.093136e+10] = 28814991000.0
Z0h[np.isnan(Z0h)] = 1e-38
show(Z0h)
Upbl = u_pbl(u, 2, z0m, z_pbl)
upbl = Upbl[0]
show(upbl)
print("Calculating Dry Limit...")
H_d = np.copy(RN)
FRUstar_i = FRUstar(z_pbl, upbl, hst, z0m, Z0h, Alt, RH, TaK, LST)
RUstar = FRUstar_i[0]
show(RUstar)
RL = FRUstar_i[1]
show(RL)
kB = KB_1(u, RUstar, z0m, Fc, LAI, TaK, pa, h)
show(kB)
Z0h = z0h(kB, z0m)
show(Z0h)
print("Calculating Wet Limit...")
# For completely wet areas
# Wet-limit stability length
L_e = 2.430E+06  # Latent heat of vapourization (J kg-1) (Brutsaert, 1982)
L_w = (np.nanmin(RUstar) ** 3.0) * rhoam / (0.61 * k * g * RN / L_e)
# C_wet = ifthenelse(z_pbl >= hst, Cw(z_pbl, L_w, z0m, z0h), PSIh_y(pcrumin(z_pbl/L_w)))
# Initial conditional mask
Mask_Cwet1 = np.copy(z_pbl)
Mask_Cwet1[Mask_Cwet1 >= hst] = -9999.0
Mask_Cwet1[Mask_Cwet1 != 9999.0] = 0
Mask_Cwet1[Mask_Cwet1 == 9999.0] = 1
C_wet1 = Cw(z_pbl, L_w, z0m, Z0h)
C_wet1 = C_wet1 * Mask_Cwet1
C_wet1 = np.nan_to_num(C_wet1)
# Second conditional mask
Mask_Cwet2 = np.copy(z0)
Mask_Cwet2[Mask_Cwet2 >= hst] = -9999.0
Mask_Cwet2[Mask_Cwet2 != 9999.0] = 1
Mask_Cwet2[Mask_Cwet2 == 9999.0] = 0
C_wet2 = np.nanmin(PSIh_y(z_pbl / L_w))
C_wet2 = C_wet2 * Mask_Cwet2
C_wet2 = np.nan_to_num(C_wet2)
# SUM CONDITIONAL
C_wet = C_wet1 + C_wet2
zd0 = z_pbl - d
zdm = np.log(zd0 / z0m)
zdh = np.log(zd0 / Z0h)
zdh[zdh == np.inf] = 375579840.0  # high resistance
zdh[np.isnan(zdh)] = 375579840
# Wet-limit external resistance
re_w = (zdh - C_wet) / (k * RUstar)
# re_w = ifthenelse(re_w <= 0.0, zdh / (k * RUstar), re_w)
# Initial conditional mask
Mask_re_w1 = np.copy(re_w)
Mask_re_w1[Mask_re_w1 <= 0.0] = -9999.0
Mask_re_w1[Mask_re_w1 != -9999.0] = 0
Mask_re_w1[Mask_re_w1 == -9999.0] = 1
re_w1 = zdh / (k * RUstar)
re_w1[re_w1 == -np.inf] = 0
re_w1[re_w1 == np.inf] = 0
re_w1 = re_w1 * Mask_re_w1
re_w1 = np.nan_to_num(re_w1)
# Second conditional mask
Mask_re_w2 = np.copy(re_w)
Mask_re_w2[Mask_re_w2 <= 0.0] = -9999.0
Mask_re_w2[Mask_re_w2 != -9999.0] = 1
Mask_re_w2[Mask_re_w2 == -9999.0] = 0
re_w2 = re_w
re_w2[re_w2 == -np.inf] = 0
re_w2[re_w2 == np.inf] = 0
re_w2 = re_w2 * Mask_re_w2
re_w2 = np.nan_to_num(re_w2)
# SUM CONDITIONAL
rew = re_w1 + re_w2
rew[rew == 0] = np.nanmax(rew)
# Wet-limit heat flux
#slopef = 17.502 * 240.97 * ea / (TaK + 240.97) ** 2.0 #SEBS
slopef = (4098 * (0.6018 * np.exp((17.27 * (TaK - 273.15)) / ((TaK - 273.15) + 237.7))) / ((TaK - 273.15) + 237.7) ** 2) * 1000  # Slope of the saturation of vapor pressure [Pa/Â°C]: divided by 1000 get in Pascal
#slopef = (4098*ea/((TaK-273.15)+237.7)**2)
gamma = (((1.013 * 10 ** -3) * pa) / (ee * (L_e / 10 ** 6))) * 1000  # Psychrometric Constant [Pa/Â°C]:divided by 1000 to get in Pascal #FAO56
#gamma = 67.0   # psychrometric constant (Pa K-1)
rhoa = 1.23  # surface air density
rhoacp = rhoa * Cp
H_w = (RN - (rhoacp / rew) * ((ea - es) / gamma)) / (1.0 + slopef / gamma)
H_w[H_w < 0] = 0
LEwet = RN - H_w
# Sensible Heat flux
print("Calculating sensible heat flux...")
# C_i = ifthenelse(z_pbl >= hst, Cw(z_pbl, RL, z0m, z0h), PSIh_y(pcrumin(z_pbl)/RL))
# Initial conditional mask
RUstar[np.isnan(RUstar)] = np.nanmin(RUstar)
RL[np.isnan(RL)] = np.nanmin(RL)
# BAS MODELATION
Mask_C_i1 = np.copy(z_pbl)
Mask_C_i1[Mask_C_i1 >= hst] = -9999.0
Mask_C_i1[Mask_C_i1 != 9999.0] = 0
Mask_C_i1[Mask_C_i1 == 9999.0] = 1
C_i_1 = Cw(z_pbl, RL, z0m, Z0h)
C_i_1 = C_i_1 * Mask_C_i1
C_i_1 = np.nan_to_num(C_i_1)
# MOS MODELATION
# Second conditional mask
Mask_C_i2 = np.copy(z_pbl)
Mask_C_i2[Mask_C_i2 >= hst] = -9999.0
Mask_C_i2[Mask_C_i2 != 9999.0] = 1
Mask_C_i2[Mask_C_i2 == 9999.0] = 0
C_i_2 = PSIh_y(z_pbl / RL)
C_i_2 = C_i_2 * Mask_C_i2
C_i_2 = np.nan_to_num(C_i_2)
# SUM CONDITIONAL OF PIXELS OF MOS & BAS
C_i = C_i_1 + C_i_2
# external resistance
re_i = (zdh - C_i) / (k * RUstar)
"""T0Ta calculation"""
t_c = np.log((z_pbl - d) / Z0h) / np.log((alt_ms - d) / Z0h)
t_pbl_A = LST * (1.0 - t_c) + TaK * t_c
t_pbl_A = t_pbl_A / (1.0 - Alt / 44331.0) ** 1.5029
t_pbl = t_pbl_A
helpvar1 = Alt / 44331.0
helpvar2 = 1.0 - helpvar1
T0 = LST / helpvar2 ** 1.5029
z_pbl_A = 1000.0  # PBL height [m]
p_pbl_A = p_s * ((44331.0 - (Alt + z_pbl_A)) / (44331.0 - alt_ms)) ** (1.0 / 0.1903)
#Theta_a = t_pbl * (101325 / p_pbl_A) ** 0.286  # provo ad utilizzare questo Theta_a, ma mi vengono valori troppo elevati!
Theta_a = (TaK - 0.6 * 20) * (101325 / p_s) ** 0.286 #questo è quello di Italo ma non so perchè è così
show(Theta_a)
Theta_s = T0
T0ta = Theta_s - Theta_a
#T0ta = lst_b1 - TaK #questa è un'approssimazione
"""End calculation"""
H_i = rhoacp * (T0ta) / re_i
# H_i = ifthenelse(H_i > H_d, H_d, H_i)
print('Percentile 10 H: ' + str(np.nanpercentile(H_i, 10)))
print('Percentile 50 H: ' + str(np.nanpercentile(H_i, 50)))
print('Percentile 90 H: ' + str(np.nanpercentile(H_i, 90)))
# H_i = ifthenelse(H_i < H_w, H_w, H_i)
# report(H_i, hmap)
# Calculate evaporation variables
print("Calculating relative evaporation and evaporative fraction...")
# Calculate relative evaporation
# Ef =(EfR*LETwet)/(RN)
ev_r = 1.0 - (H_i - H_w) / (H_d - H_w)  # set water and wet surfaces to 1.0
# report(ev_r, evaprmap)
# Calculate evaporative fraction
ETF = ev_r * (1.0 - H_w / H_d)
# report(Evapfr, evapfrmap)
Lambda = 2501000 - (2361 * (TaK - 273.15))  # Latent heat of vaporization [J]
# Calculate latent energy flux
print("Calculating Latent Energy Flux...")
labdaE = ETF * (RN)  # Latent Heat Flux [W/m2] (Su, 2002)
labdaE2 = RN - H_i  # Latent Heat Flux [W/m2] (Su, 2002)
ET_Instant = RN - H_i - my_G0 #[W/m2]
# Evapfr = labdaE2/RN
ETR = (labdaE2 / Lambda) * Cdi * (24 * 3600)  # Actual Evapotranspiration [mm/day]
ETR[ETR > 12] = 12
ETR[ETR < 0] = 0
ETR = ETR
ETF[ETF > 1] = 1
ETF[ETF < 0] = 0
print('Actual Evapotranspiration average:' + str(np.nanmean(ETR)), 'mm/day')
print('Actual Evapotranspiration variability:' + str(np.nanstd(ETR)), 'mm/day')
show(ETR)


#Actual Evapotranspiration average:2.292401851975694 mm/day
#Actual Evapotranspiration variability:2.4071460869212644 mm/day

